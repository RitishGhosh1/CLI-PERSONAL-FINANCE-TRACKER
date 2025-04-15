import json
import re
import csv
import sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from tabulate import tabulate

def main():
    transactions_file = "transactions.json"
    transactions = load_data(transactions_file)
    print("\nPersonal Finance Tracker")
    username=input("Enter your Username ")
    email=input("Enter your email ")
    emailvalidity=r"([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+"
    if re.search(emailvalidity,email):
        while True:
            print("1. Add Transaction")
            print("2. Remove Transaction")
            print("3. View Summary")
            print("4. Export to Excel")
            print("5. Export to CSV")
            print("6. Exit")
            choice = input("Choose an option: ")
            if choice == "1":
                add_transaction(transactions, transactions_file)
            elif choice == "2":
                category = input("Enter category to filter (or press Enter for all): ").capitalize()
                remove_transaction(transactions,category)
            elif choice=="3":
                category=input("Enter category to filter (or press Enter for all): ").capitalize()
                summary(transactions,category if category else None)
            elif choice == "4":
                exporttoxls(transactions,username,email)
            elif choice=="5":
                exporttocsv(transactions)
            elif choice == "6":
                sys.exit("Goodbye!")
            else:
                print("Invalid choice. Try again.")
    else:
        print("Invalid Email")

    
def add_transaction(transactions,file):
    try:
        type_=input("Type (income or expense): ").lower()
        if type_ not in ['income','expense']:
            raise ValueError("Type must be in 'income' or 'expense'")
        category=input('Category (e.g Food, Salary, Grocery etc):')
        amount=float(input("Amount: "))
        if amount<=0:
            raise ValueError("Amount must be positive")
        description=input("Description: ")
        Date=datetime.now().strftime("%Y-%m-%d")
        transaction={
            'date':Date,
            'category':category,
            'description':description,
            'type':type_,
            'amount':amount
        }
        transactions.append(transaction)
        with open(file,'w') as file:
            json.dump(transactions, file, indent=4)
        print("Transaction Added")
    except ValueError as e:
        print(f"Error: {e}")
    
def summary(transactions, category=None):
    total_income=0
    expense=0
    table=[]
    for t in transactions:
        if t['type'] == 'income':
            total_income += t['amount']
    for t in transactions:
        if category == None or t['category'] == category:
            if t['type'] == 'expense':
                expense += t['amount']
            table.append([t['date'], t['category'], t['description'], t['type'].capitalize(), t['amount']])
    balance = total_income - expense
    if table:
        print(tabulate(table,headers=["Date","Category","Description","Type","Amount"],tablefmt="grid"))
        print(f"Total income: {total_income:0.2f}")
        print(f"Total Expense: {expense:0.2f}")
        print(f"Total Balance: {balance:0.2f}")
    else:
        print(f"No Transaction Record found")

def remove_transaction(transactions,category):
    found=False
    ta=[]
    for t in transactions:
        if t['category'] == category:
            ta.append([t['date'], t['category'], t['description'], t['type'].capitalize(), t['amount']])
            transactions.remove(t)
            found=True
    if found:
        print("\n Transaction Removed ")
        for transactions in ta:
            print(transactions)
    else:
        print("Transaction not found")
            
def exporttoxls(transactions,username,email,filename='transactions.xlsx'):
    wb=Workbook()
    ws=wb.active
    ws.title='Transaction'
    bold_font=Font(bold=True)
    ws['A1']=f"Username"
    ws['B1']=f"{username}"
    ws['A2']=f"Email"
    ws['B2']=f"{email}"
    ws['A1'].font=bold_font
    ws['B1'].font=bold_font
    ws['A2'].font=bold_font
    ws['B2'].font=bold_font
    headers=['Date','Category','Description','Type','Amount']
    ws.append([])
    ws.append(headers)
    for col in ws[4]:
        col.font=bold_font
    for t in transactions:
        ws.append([t['date'],t['category'],t['description'],t['type'],t['amount']])
    wb.save(filename)
    print(f"File is exported") 

def exporttocsv(transactions,filename='transactions.csv'):
    headers=['date','category','description','type','amount']
    with open(filename,'w',newline='') as file:
        writer=csv.DictWriter(file,fieldnames=headers)
        writer.writeheader()
        writer.writerows(transactions)

def load_data(filename):
    try:
        with open(filename, "r") as file:
            return json.load(file)
    except (FileNotFoundError, json.JSONDecodeError):
        return []
    
if __name__=="__main__":
    main()
        